"""Import des données réelles de production (fichiers Excel) dans la base RestoConnect.

Usage (depuis backend/, venv activé) :
    python -m app.database.import_prod --data-dir /tmp/restoducoeur/restoducoeur
    python -m app.database.import_prod --data-dir ... --force

Idempotent par défaut : skip si la table cible est déjà remplie.
`--force` vide les tables cibles avant import (ordre inverse des FK).
"""

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import bcrypt
import openpyxl
from sqlalchemy import text

from app.database.connection import SessionLocal, engine
from app.database.models import Center, Stock, User, Vehicule
from app.enums import (
    CenterStatus,
    StockCategory,
    StockStatus,
    UserStatus,
    VehiculeCategory,
    VehiculeStatus,
)

# ---------------------------------------------------------------------------
# Chemins des fichiers Excel (relatifs à --data-dir)
# ---------------------------------------------------------------------------
CENTRES_FILE = (
    "Cellule Informatique/1. INFORMATIONS ASSOCIATION/1. Centres Restos du coeur/"
    "Centres Restos du coeur de Seine et Marne.xlsx"
)
VEHICULES_FILE = "Copie de Base véhicules 08 01 25 PYR.xlsx"
MATERIELS_FILE = (
    "Cellule Informatique/5. PROJETS IT/2. BUILD/2. PROJETS OLD/1. INVENTAIRE DES STOCKS/"
    "2. APPLICATION MOBILE (old_PA)/3. CONCEPTION/2. LIVRABLE DE PREPARATION/"
    "2. Tableau Excel Inventaire/3. Archive/Inventaire RDC – Copie.xlsx"
)


# ---------------------------------------------------------------------------
# Helpers de normalisation
# ---------------------------------------------------------------------------
def _strip_accents(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _norm(s):
    """Normalise pour comparaison : uppercase, sans accents, espaces multiples -> un."""
    if s is None:
        return None
    s = _strip_accents(str(s)).upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def _norm_key(s):
    """Normalisation agressive pour la résolution : retire tirets, apostrophes, espaces."""
    if s is None:
        return None
    s = _strip_accents(str(s)).upper()
    s = re.sub(r"[\s\-'’_]", "", s)
    return s or None


def _clean(s):
    if s is None:
        return None
    s = str(s).strip()
    return s or None


def _to_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


# ---------------------------------------------------------------------------
# Mapping familles matériel -> StockCategory
# ---------------------------------------------------------------------------
FAMILLE_TO_CATEGORY = {
    "INFORMATIQUE": StockCategory.INFORMATIQUE,
    "TELEPHONE": StockCategory.INFORMATIQUE,
    "CONGELATEUR": StockCategory.REFRIGIRE,
    "REFRIGERATEUR": StockCategory.REFRIGIRE,
    "ARMOIRE FRIGO": StockCategory.REFRIGIRE,
    "MATERIELS DE CUISINE": StockCategory.RESTAURATION,
    "BALANCE": StockCategory.RESTAURATION,
    "BAC": StockCategory.RESTAURATION,
    "MATERIEL DE BUREAU": StockCategory.BUREAU,
    "MOBILIER DE BUREAU": StockCategory.BUREAU,
    "RAYONNAGE": StockCategory.BUREAU,
    "ENTRETIEN": StockCategory.OTHER,
    "MANUTENTION": StockCategory.OTHER,
}


def map_famille(famille):
    key = _norm(famille)
    if key is None:
        return StockCategory.OTHER
    return FAMILLE_TO_CATEGORY.get(key, StockCategory.OTHER)


# ---------------------------------------------------------------------------
# Mapping type véhicule -> VehiculeCategory
# ---------------------------------------------------------------------------
def map_vehicule_category(type_, modele):
    t = _norm(type_) or ""
    m = _norm(modele) or ""
    if "FRIGO" in t or "FRIGO" in m:
        return VehiculeCategory.FRIGORIFIQUE
    if "BUS" in t or "BUS" in m:
        return VehiculeCategory.CAMION
    if "HAYON" in t:
        return VehiculeCategory.FOURGON
    if t.startswith("VP") or any(
        k in m for k in ("KANGOO", "CLIO", "MEGANE", "EXPRESS", "PARTNER", "COMBO")
    ):
        return VehiculeCategory.VOITURE
    if t.startswith("UL"):
        return VehiculeCategory.UTILITAIRE
    return VehiculeCategory.UTILITAIRE


def map_vehicule_status(etat):
    e = _norm(etat) or ""
    if "HS" in e or "FIN DE VIE" in e:
        return VehiculeStatus.OUT_OF_SERVICE
    if "STAND BY" in e or "REVISION" in e or "MAINTENANCE" in e:
        return VehiculeStatus.IN_MAINTENANCE
    if "REPARATION" in e:
        return VehiculeStatus.UNDER_REPAIR
    return VehiculeStatus.IN_SERVICE


# ---------------------------------------------------------------------------
# Résolution des centres
# ---------------------------------------------------------------------------
# Alias : clé normalisée (sans espaces/tirets) -> nom canonique de centre
ALIAS = {
    "SIEGE": "Siège",
    "SIEGEREMPLACEMENT": "Siège",
    "TOITSDUCOEURSIEGE": "Siège",
    "STOCKSIEGE": "Siège",
    "ECUELLES": "Centre d'ECUELLES",
    "MELUNLAVOISIER": "Centre de MELUN-LAVOISIER",
    "MELUNALMONT": "Centre de MELUN-ALMONT",
    "CENTREDEMELUNALMONT": "Centre de MELUN-ALMONT",
    "CENTRESMELUNETDAMMARIE": "Centre de MELUN-ALMONT",
    "CENTREDESTTHIBAULTDESVIGNESTORCY": "Centre de TORCY",
    "VERTSAINTDENIS": "Centre de VERT-SAINT-DENIS",
    "LAFERTE": "Centre de LA FERTE SOUS JOUARRE",
    "LEMEE": "Centre de LE-MEE-SUR-SEINE",
    "LIZYSUROURQ": "Centre de LIZY-SUR-OURCQ",
    "LIZYSUROURCQ": "Centre de LIZY-SUR-OURCQ",
    "STMARD": "Centre de SAINT-MARD",
    "CENTREDESTMARD": "Centre de SAINT-MARD",
    "CENTREESBLY": "Centre d'ESBLY",
    "CENTREOZOIRLAFERRIERE": "Centre d'OZOIR-LA-FERRIERE",
    "BRAY": "Centre de BRAY-SUR-SEINE",
    "BRIE": "Centre de BRIE-COMTE-ROBERT",
    "CHAMPS": "Centre de CHAMPS-SUR-MARNE",
    "COMBS": "Centre de COMBS-LA-VILLE",
    "DAMMARIE": "Centre de DAMMARIE-LES-LYS",
    "MOISSY": "Centre de MOISSY-CRAMAYEL",
    "MONTEREAU": "Centre de MONTEREAU-FAULT-YONNE",
    "OZOIR": "Centre d'OZOIR-LA-FERRIERE",
    "PONTAULT": "Centre de PONTAULT-COMBAULT",
    "ROISSY": "Centre de ROISSY-EN-BRIE",
    "SAVIGNY": "Centre de SAVIGNY-LE-TEMPLE",
    "TOURNAN": "Centre de TOURNAN-EN-BRIE",
    "VILLEPARISIS": "Centre de VILLEPARISIS",
    "AVON": "Centre d'AVON",
    "ESBLY": "Centre d'ESBLY",
    "MEAUX": "Centre de MEAUX",
    "LAGNY": "Centre de LAGNY-SUR-MARNE",
    "NEMOURS": "Centre de NEMOURS",
    "NANGIS": "Centre de NANGIS",
    "PROVINS": "Centre de PROVINS",
    "TORCY": "Centre de TORCY",
    "COULOMMIERS": "Centre de COULOMMIERS",
    "CHAMPAGNE": "Centre de CHAMPAGNE-SUR-SEINE",
    # Entrepôts / dépôts / services (créés comme entrepôts)
    "ENTREPOTDUNORDQUINCY": "Entrepôt Nord Quincy",
    "ENTREPOTNORDQUINCY": "Entrepôt Nord Quincy",
    "ENTREPOTSUDSAVIGNYLETEMPLE": "Entrepôt Sud Savigny",
    "ENTREPOTSUDSAVIGNY": "Entrepôt Sud Savigny",
    "DEPOTQUINCYVOISINS": "Dépôt Quincy Voisins",
    "DEPOTSAVIGNY": "Dépôt Savigny",
    "DEPOTSAVIGNYSTOCK": "Dépôt Savigny",
    "DEPOTVERNEUIL": "Dépôt Verneuil",
    "EIF": "EIF",
    "BUSDUCOEUR": "Bus du Cœur",
    "FORMATIONNORD": "Formation Nord",
}

WAREHOUSE_NAMES = {
    "Entrepôt Nord Quincy",
    "Entrepôt Sud Savigny",
    "Dépôt Quincy Voisins",
    "Dépôt Savigny",
    "Dépôt Verneuil",
    "EIF",
    "Bus du Cœur",
    "Formation Nord",
}


def _clean_center_name(nom):
    name = re.sub(r"^AD77\s*[–-]\s*", "", nom)
    name = re.sub(r"^Christian CHARLUET \(Siège\)$", "Siège", name, flags=re.IGNORECASE)
    return name.strip()


def import_centres(data_dir: Path, db) -> dict:
    path = data_dir / CENTRES_FILE
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Feuil1"]

    name_to_id = {}
    created = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        nom = _clean(row[0])
        if not nom:
            continue
        name = _clean_center_name(nom)
        center = Center(
            name=name,
            street=_clean(row[1]),
            city=_clean(row[3]),
            postal_code=_clean(row[2]),
            telephone=_clean(row[4]),
            email=_clean(row[5]),
            status=CenterStatus.OPEN,
            is_warehouse=False,
        )
        db.add(center)
        db.flush()
        name_to_id[_norm_key(name)] = center.id
        created += 1

    db.commit()
    print(f"[centres] {created} importés")
    return name_to_id


def resolve_center(location, name_to_id, db):
    """Résout une chaîne d'affectation/centre vers un center_id (crée un entrepôt si besoin)."""
    if not location:
        return None
    key = _norm_key(location)
    if not key:
        return None

    # 1. alias explicite (clé normalisée)
    alias_key = _norm_key(location)
    if alias_key in ALIAS:
        canonical = ALIAS[alias_key]
        cid = name_to_id.get(_norm_key(canonical))
        if cid is not None:
            return cid
        is_wh = canonical in WAREHOUSE_NAMES
        center = Center(
            name=canonical,
            city="Seine-et-Marne",
            status=CenterStatus.OPEN,
            is_warehouse=is_wh,
        )
        db.add(center)
        db.flush()
        name_to_id[_norm_key(canonical)] = center.id
        return center.id

    # 2. correspondance directe
    cid = name_to_id.get(key)
    if cid is not None:
        return cid

    # 3. correspondance par sous-chaîne
    for cname, cid in list(name_to_id.items()):
        if cname and (cname in key or key in cname):
            return cid

    # 4. fallback : créer un entrepôt
    center = Center(
        name=location.strip(),
        city="Seine-et-Marne",
        status=CenterStatus.OPEN,
        is_warehouse=True,
    )
    db.add(center)
    db.flush()
    name_to_id[key] = center.id
    return center.id


# ---------------------------------------------------------------------------
# Import véhicules
# ---------------------------------------------------------------------------
def import_vehicules(data_dir: Path, db, name_to_id: dict) -> int:
    path = data_dir / VEHICULES_FILE
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["liste vehicules"]

    created = 0
    seen_immat = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        immat = _clean(row[0])
        if not immat or immat in seen_immat:
            continue
        seen_immat.add(immat)

        marque = _clean(row[1])
        modele = _clean(row[2])
        type_ = _clean(row[3])
        date_ct = _to_date(row[15])
        klm = _to_int(row[25])
        affectation = _clean(row[20])
        etat = _clean(row[22])

        name = f"{marque} {modele}".strip() if marque else immat
        center_id = resolve_center(affectation, name_to_id, db)

        vehicule = Vehicule(
            name=name,
            immatriculation=immat,
            category=map_vehicule_category(type_, modele),
            status=map_vehicule_status(etat),
            nb_km=klm or 0,
            last_technical_inspection_date=date_ct,
            next_technical_inspection_date=date_ct,
            center_id=center_id,
        )
        db.add(vehicule)
        created += 1

    db.commit()
    print(f"[vehicules] {created} importés")
    return created


# ---------------------------------------------------------------------------
# Import matériels
# ---------------------------------------------------------------------------
def import_materiels(data_dir: Path, db, name_to_id: dict) -> int:
    path = data_dir / MATERIELS_FILE
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Inventaire au 290125"]

    created = 0
    seen_ref = set()
    auto_ref = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Colonnes réelles (le header Excel est décalé d'une colonne) :
        # [1] CODE ARTICLE, [2] LIBELLE, [3] FAMILLE, [4] FOURNISSEUR,
        # [5] DATE ACHAT, [6] VALEUR ACHAT, [13] CENTRE, [14] LOCALISATION
        code = _clean(row[1])
        libelle = _clean(row[2])
        famille = _clean(row[3])
        valeur = _to_int(row[6])
        date_achat = _to_date(row[5])
        centre = _clean(row[13])

        if not libelle and not code:
            continue
        if libelle and _norm(libelle) in ("TOTAL", "SOUS-TOTAL", "SOUS TOTAL"):
            continue

        reference = code
        if not reference or reference.upper() == "NON IMAT":
            auto_ref += 1
            reference = f"AUTO-{auto_ref:05d}"
        if reference in seen_ref:
            continue
        seen_ref.add(reference)

        center_id = resolve_center(centre, name_to_id, db)

        stock = Stock(
            name=libelle or reference,
            category=map_famille(famille),
            reference=reference,
            qr_code=reference,
            status=StockStatus.AVAILABLE,
            creation_date=date_achat,
            last_scan_date=date_achat,
            rating=valeur,
            center_id=center_id,
        )
        db.add(stock)
        created += 1

    db.commit()
    print(f"[materiels] {created} importés")
    return created


# ---------------------------------------------------------------------------
# Import users (comptes de démonstration, rattachés aux centres réels)
# ---------------------------------------------------------------------------
def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def import_users(db, name_to_id: dict) -> int:
    """Crée les comptes de démo, rattachés aux centres réels importés.

    Le superadmin/admin sont rattachés au Siège ; les responsables de centre
    aux premiers centres réels disponibles.
    """
    centre_ids = sorted(set(name_to_id.values()))
    siege_id = name_to_id.get(_norm_key("Siège"))
    # centres non-entrepôts en priorité, hors Siège (centre administratif)
    real_centres = [
        cid
        for cid in centre_ids
        if cid != siege_id
        and (c := db.get(Center, cid)) is not None
        and not c.is_warehouse
    ]

    def cid_or(i):
        return (
            real_centres[i] if i < len(real_centres) else (siege_id or real_centres[0])
        )

    def center_fields(cid):
        """Dérive city/postal_code/street du centre rattaché (le profil les affiche)."""
        c = db.get(Center, cid)
        if c is None:
            return {}, {}, {}
        return (
            {"city": c.city},
            {"postal_code": c.postal_code},
            {"street": c.street},
        )

    def make_user(name, lastname, email, status, cid):
        city, postal, street = center_fields(cid)
        return User(
            name=name,
            lastname=lastname,
            email=email,
            password=_hash_password("1234"),
            telephone="0123456789",
            status=status,
            center_id=cid,
            created_at=date.today(),
            updated_at=date.today(),
            **city,
            **postal,
            **street,
        )

    users = [
        make_user(
            "Antoine",
            "Lefebvre",
            "superadmin@resto.com",
            UserStatus.SUPER_ADMIN,
            cid_or(0),
        ),
        make_user("Julie", "Moreau", "admin@resto.com", UserStatus.ADMIN, cid_or(0)),
        make_user(
            "Marc", "Dubois", "resp1@resto.com", UserStatus.CENTER_ADMIN, cid_or(0)
        ),
        make_user(
            "Sophie", "Bernard", "resp2@resto.com", UserStatus.CENTER_ADMIN, cid_or(1)
        ),
        make_user(
            "Karim",
            "Benali",
            "vehicule1@resto.com",
            UserStatus.VEHICULE_ADMIN,
            cid_or(0),
        ),
        make_user(
            "Hugo", "Petit", "stock1@resto.com", UserStatus.STOCK_ADMIN, cid_or(0)
        ),
        make_user("Paul", "Fontaine", "user@resto.com", UserStatus.User, cid_or(0)),
    ]
    db.add_all(users)
    db.commit()
    print(f"[users] {len(users)} importés")
    return len(users)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--data-dir", required=True, help="Racine extraite de l'archive"
    )
    parser.add_argument(
        "--force", action="store_true", help="Vider les tables cibles avant import"
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    if not data_dir.exists():
        raise SystemExit(f"data-dir introuvable : {data_dir}")

    db = SessionLocal()

    if args.force:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "TRUNCATE stock_events, inventory_stock, inventory, "
                    "vehicule_documents, vehicule, stock, center CASCADE"
                )
            )
        print("[force] tables vidées")

    if db.query(Center).first():
        print("Centres déjà présents, skip (utilisez --force pour réimporter)")
        db.close()
        return

    name_to_id = import_centres(data_dir, db)
    import_vehicules(data_dir, db, name_to_id)
    import_materiels(data_dir, db, name_to_id)
    import_users(db, name_to_id)

    db.close()
    print("Import terminé ✅")


if __name__ == "__main__":
    main()
